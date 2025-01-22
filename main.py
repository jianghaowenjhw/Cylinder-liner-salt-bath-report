# 此文件用来给盐浴检查报告汇总文件的工作提供便利, 非相关者请勿使用
# 作者: zhugeshu
# 创作日期: 2025年1月21日
# 最后修改日期: 2025年1月22日
# 主要功能介绍: 0. 读取往期盐浴实验室报告, 并写入csv文件; 1. 标准录入, 录入实验室报告并写入csv文件; 2. 信息检查与清洗; 3. 查询信息; 4. 接收需要的编号列表, 生成验收报告

import csv
import openpyxl
from datetime import datetime, timedelta
import os
import time
from copy import copy

class Product:
    def __init__(
        self, 
        year: int, # 检查的年份, 以整数储存
        month: int, # 检查的月份, 以整数储存
        day: int, # 检查的日期, 以整数储存
        heat_number: int, # 以整数储存, 但应当有前导0, 共4位
        compound_layer: str, # 以字符串格式保存, 保证精度
        diffusion_depth: str, # 以字符串格式保存, 保证精度
        preid: int, # 生产的年份, 以整数储存
        identification: int # 产品的标识编号, 以整数储存, 但应当有前导0, 共5位
    ) -> None:
        self.year = int(year)
        self.month = int(month)
        self.day = int(day)
        self.heat_number = int(heat_number)
        self.compound_layer = str(compound_layer)
        self.diffusion_depth = str(diffusion_depth)
        self.preid = int(preid)
        self.identification = int(identification)

def get_project_root() -> str:
    """
    获取项目的根目录。
    """
    return os.path.dirname(os.path.abspath(__file__))

def get_absolute_path(relative_path: str) -> str:
    """
    将相对路径转换为绝对路径。
    """
    return os.path.join(get_project_root(), relative_path)

def get_relative_path(absolute_path: str) -> str:
    """
    获取相对于项目根目录的路径。
    """
    return os.path.relpath(absolute_path, get_project_root())

def read_cell(
    file_path: str, 
    sheet_name: str, 
    cell: tuple[int, int]
):
    """
    读取 Excel 文件中指定单元格的内容。
    """
    # 打开 Excel 文件
    file_path = get_absolute_path(file_path)
    workbook = openpyxl.load_workbook(file_path)
    # 选择工作表
    sheet = workbook[sheet_name]
    # 读取指定单元格的内容
    cell_value = sheet.cell(row=cell[0], column=cell[1]).value
    # 关闭 Excel 文件
    workbook.close()
    return cell_value

def excel_number_to_date(
    excel_number: int
) -> datetime:
    """
    将Excel中的日期序列号转换为实际日期。
    """
    '''
    举例:
    excel_number = 45674
    date = excel_number_to_date(excel_number)
    print(date.strftime("%Y/%m/%d"))
    '''
    excel_base_date = datetime(1899, 12, 30)  # Excel的起始日期是1899年12月30日（第0天）
    delta = timedelta(days=excel_number)
    actual_date = excel_base_date + delta
    return actual_date

def get_date_components(
    excel_number: int
) -> tuple[int, int, int]:
    """
    将Excel日期序列号转换为实际日期，并分别提取年份、月份和日期。
    """
    date = excel_number_to_date(int(excel_number))
    year = date.year
    month = date.month
    day = date.day
    return year, month, day

def date_components_to_excel_number(year: int, month: int, day: int) -> int:
    """
    将年份、月份和日期转换为Excel日期序列号。
    """
    excel_base_date = datetime(1899, 12, 30)  # Excel的起始日期是1899年12月30日（第0天）
    date = datetime(year, month, day)
    delta = date - excel_base_date
    return delta.days

def is_product_in_csv(
    product: Product, 
    file_name: str
) -> bool:
    """
    检查产品是否在指定的 CSV 文件中。
    """
    file_name = get_absolute_path(file_name)
    if not os.path.exists(file_name):
        print(f"文件 {file_name} 不存在。")
        return False

    with open(file_name, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头
        for row in reader:
            if (
                int(row[0]) == product.year and
                int(row[1]) == product.month and
                int(row[2]) == product.day and
                int(row[3]) == product.heat_number and
                str(row[4]) == product.compound_layer and
                str(row[5]) == product.diffusion_depth and
                int(row[6]) == product.preid and
                int(row[7]) == product.identification
            ):
                return True
    return False

def write_into_csv(
    products: list[list[Product]]
) -> None:
    """
    将产品信息写入 csv 文件。
    每行依次是：年份、月份、日期、热处理号、复合层、扩散层、标识。
    注意, products是一个二维列表, 每个一维列表中的product除了标识外, 其他信息都是相同的。
    注意, 25年的产品需要写入"25database.csv", 24年的产品需要写入"24database.csv", 以此类推。
    """
    if len(products) == 0:
        print("写入不成功")
        return

    product_dict = {}
    
    for product_list in products:
        for product in product_list:
            year = product.preid
            if year not in product_dict:
                product_dict[year] = []
            product_dict[year].append(product)
    
    for preid, product_list in product_dict.items():
        file_name = f"{preid}database.csv"
        file_name = get_absolute_path(file_name)
        file_exists = os.path.exists(file_name)
        
        with open(file_name, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            if not file_exists:
                writer.writerow(["year", "month", "day", "heat_number", "compound_layer", "diffusion_depth", "preid", "identification"])
            
            for product in product_list:
                if is_product_in_csv(product, file_name):
                    print(f"产品 {product.identification} 已存在于 {file_name} 中。")
                else:
                    writer.writerow([
                        product.year,
                        product.month,
                        product.day,
                        product.heat_number,
                        product.compound_layer,
                        product.diffusion_depth,
                        product.preid,
                        product.identification
                    ])
    print("产品信息已成功写入 csv 文件。")

def search_product_in_csv(
    full_identification: str
) -> tuple[int, int, int, int, float, float, int, int]:
    '''
    参数为完整的identification, 查询对应的数据库, 存在的话打印对应的信息
    '''
    (preid, identification) = clear_abnormal_characters(full_identification).split()
    file_name = f"{preid}database.csv"
    file_name = get_absolute_path(file_name)
    if not os.path.exists(file_name):
        print(f"文件 {file_name} 不存在。")
        return
    
    with open(file_name, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头
        for row in reader:
            if int(row[7]) == int(identification):
                print(
                    f"在{preid}database.csv文件中, 年份: {row[0]}, 月份: {row[1]}, 日期: {row[2]}, 热处理号: {row[3]}, 复合层: {row[4]}, 扩散层: {row[5]}, preid: {row[6]}, identification: {row[7]}"
                )
            
    return

def check_just(
    products: list[list[Product]]
) -> None:
    '''
    读取之后立刻检查是否有问题
    '''
    for product_list in products:
        for product in product_list:
            if (
                product.year == None or
                product.month == None or
                product.day == None or
                product.heat_number == None or
                product.compound_layer == None or
                product.diffusion_depth == None or
                product.identification == None or
                int(product.identification) >= 100000
            ):
                print("读取的数据有问题, 你自己看看怎么回事")
                products = []
    pass

def clear_abnormal_characters(
    identification: str
) -> str:
    '''
    把identification中的除数字和空格外的所有字符变成空格, 并把连续的空格变成一个空格, 最后清除首尾的空格
    '''
    result = ""
    # 非数字和空格的字符变成空格
    for char in identification:
        if char.isdigit() or char.isspace():
            result += char
        else:
            result += " "
    # 连续的空格变成一个空格, 清除首尾的空格
    result = " ".join(result.split())
    return result

def get_ID_with_preid(
    response: str
) -> list[list[int]]:
    '''
    通过输入的字符串, 获取对应的编号列表
    输入形如"24 07698 13456 25 19778", 其中24和25表示年份, 由preid承接
    输出形如[[24, 7698], [24, 13456], [25, 19778]], 其中每个子列表的第一个元素是preid, 第二个元素是identification
    '''
    response = response.split()
    list_of_identification = []
    preid = time.localtime()[0] - 2000
    for i in range(len(response)):
        if int(response[i]) < 50: # 50是一个随便取的数, 用来判断是否是年份
            preid = response[i]
        else:
            list_of_identification.append([preid, response[i]])
    return list_of_identification

def get_info(
    identification_with_preid: list[int]
) -> tuple[int, int, int, int, str, str, int, int]:
    '''
    通过identification_with_preid, 查询对应的数据库, 获取对应信息
    '''
    (preid, identification) = identification_with_preid
    file_name = f"{preid}database.csv"
    file_name = get_absolute_path(file_name)
    if not os.path.exists(file_name):
        print(f"文件 {file_name} 不存在。")
        return None
    with open(file_name, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            if int(row[7]) == int(identification):
                return (
                    int(row[0]),  # year
                    int(row[1]),  # month
                    int(row[2]),  # day
                    int(row[3]),  # heat_number
                    str(row[4]),  # compound_layer
                    str(row[5]),  # diffusion_depth
                    int(row[6]),  # preid
                    int(row[7])   # identification
                )
    print(f"数据库{preid}中没有编号为{identification}的产品")
    return None

def sort(
    table: list[Product]
) -> list[Product]:
    '''
    对table进行排序, 对于所有指标都是升序排列, 优先级为: year, mouth, day, heat_number, identification
    '''
    table.sort(key=lambda x: (x.year, x.month, x.day, x.heat_number, x.identification))
    return table

def divide(
    table: list[Product]
) -> list[list[Product]]:
    '''
    将table中的product分成多个list, 每个list中的product的year, month, day, heat_number, compound_layer, diffusion_depth都相同
    '''
    result = []
    temp = []
    for i in range(1, len(table)):
        if (table[i].year == table[i-1].year and
            table[i].month == table[i-1].month and
            table[i].day == table[i-1].day and
            table[i].heat_number == table[i-1].heat_number and
            table[i].compound_layer == table[i-1].compound_layer and
            table[i].diffusion_depth == table[i-1].diffusion_depth
            ):
            temp.append(table[i-1])
        else:
            temp.append(table[i-1])
            result.append(temp)
            temp = []
    temp.append(table[-1])
    result.append(temp)
    return result

def ready_to_xlsx(
    list_of_identification_with_preid: list[list[int]]
) -> list[list[Product]]:
    '''
    生成一个二维列表, 一维列表的元素都是product, 其中, 除了标识和preid外, 其他信息都相同的product放在同一个一维列表中
    不同的一维列表排序: 时间从旧到新, 时间相同的heat_number从小到大
    一维列表内排序: identification从小到大
    '''
    table = []
    for identification in list_of_identification_with_preid:
        (year, month, day, heat_number, compound_layer, diffusion_depth, preid, identification) = get_info(identification)
        table.append(Product(year, month, day, heat_number, compound_layer, diffusion_depth, preid, identification))
    table = sort(table)
    table = divide(table)
    return table

# 下面是一些操作xlsx文件的函数

def copy_cell(source_cell, target_cell):
    # 复制内容
    target_cell.value = source_cell.value
    # 复制格式
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

def move_cell(source_cell, target_cell):
    # 复制内容和格式
    copy_cell(source_cell, target_cell)
    # 清空原单元格内容和格式
    source_cell.value = None
    source_cell.font = openpyxl.styles.Font()
    source_cell.border = openpyxl.styles.Border()
    source_cell.fill = openpyxl.styles.PatternFill()
    source_cell.number_format = 'General'
    source_cell.protection = openpyxl.styles.Protection()
    source_cell.alignment = openpyxl.styles.Alignment()

def move_line(start: int, end: int, sheet: openpyxl.worksheet.worksheet.Worksheet):
    '''
    把start行的内容和格式移动到end行
    '''
    if start == end:
        return
    for i in range(1, 10):
        move_cell(sheet.cell(start, i), sheet.cell(end, i))

def copy_line(start: int, end: int, sheet: openpyxl.worksheet.worksheet.Worksheet):
    '''
    把start行的内容和格式复制到end行
    '''
    if start == end:
        return
    for i in range(1, 10):
        copy_cell(sheet.cell(start, i), sheet.cell(end, i))

def add_borders(cell):
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    cell.border = thin_border

def adjust_format(
    sheet: openpyxl.worksheet.worksheet.Worksheet, 
    line: int
) -> None:
    '''
    调整输出的检查报告的格式
    '''
    sheet.merge_cells(f"B{line+6}:C{line+6}")
    sheet.merge_cells(f"E{line+6}:H{line+6}")
    sheet.cell(2, 1).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    sheet.cell(2, 9).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")

def fill_in(
    table: list[list[Product]], 
    file_name: str
) -> None:
    '''
    将table中的信息填入file_name中
    '''
    file_name = get_absolute_path(file_name)
    if os.path.exists(file_name):
        response = input(f"文件{file_name}已存在, 是否覆盖?(y/n)")
        if response == "y" or not response:
            os.remove(file_name)
        else:
            print("操作取消")
            return
    wb = openpyxl.load_workbook("model.xlsx")
    ws = wb.active

    # 移动底层信息        
    lines = len(table)
    move_line(7, lines+6, ws)
    # 格式填充
    for i in range(len(table)):
        copy_line(6, 6+i, ws)
    # 内容填充
    for i in range(len(table)):
        num = len(table[i])
        ws.cell(6+i, 1).value = datetime(table[i][0].year, table[i][0].month, table[i][0].day)
        ws.cell(6+i, 1).number_format = "yy/m/d"
        ws.cell(6+i, 2).value = f"{table[i][0].heat_number:03d}"
        ws.cell(6+i, 3).value = num
        ws.cell(6+i, 4).value = f"{table[i][0].heat_number:03d}"  
        ws.cell(6+i, 5).value = str(table[i][0].compound_layer)
        ws.cell(6+i, 6).value = str(table[i][0].diffusion_depth)
        to_fill = ""
        for j in range(num-1):
            to_fill += f"DL{table[i][j].preid}-{table[i][j].identification:05d}  "
        to_fill += f"DL{table[i][-1].preid}-{table[i][-1].identification:05d}"
        ws.cell(6+i, 9).value = to_fill
    
    adjust_format(ws, lines)

    wb.save(file_name)

# 下面是五个主要分支函数

def read_past_reports_and_write_into_csv():
    # 读取 Excel 文件
    filepath = "total.xlsx"
    sheet = "111"
    products = []
    line = 5
    while read_cell(filepath, sheet, (line, 1)):
        products.append([])
        excel_time = read_cell(filepath, sheet, (line, 1))
        
        if isinstance(excel_time, int):
            year, month, day = get_date_components(excel_time)
        else:
            year, month, day = excel_time.split(".")
            year = int(year) + 2000
        heat_number = read_cell(filepath, sheet, (line, 2))
        conpound_layer = read_cell(filepath, sheet, (line, 5))
        diffusion_depth = read_cell(filepath, sheet, (line, 6))
        identification = read_cell(filepath, sheet, (line, 9))
        identification = clear_abnormal_characters(identification)
        identifications_with_preid = get_ID_with_preid(identification)
        for identification_with_preid in identifications_with_preid:
            products[-1].append(Product(year, month, day, heat_number, conpound_layer, diffusion_depth, identification_with_preid[0], identification_with_preid[1]))
        line += 1
    
    check_just(products)
    write_into_csv(products)

def standard_input():
    # 读取 Excel 文件
    filepath = input("请输入同目录下的文件名称:")
    sheet = input("请输入工作表名称:")
    products = []
    line = 4
    while read_cell(filepath, sheet, (line, 2)) != None:
        if not ("265" in read_cell(filepath, sheet, (line, 3))):
            line += 1
            continue
        products.append([])
        excel_time = read_cell(filepath, sheet, (line, 2))
        year, month, day = get_date_components(excel_time)
        heat_number = read_cell(filepath, sheet, (line, 4))
        conpound_layer = read_cell(filepath, sheet, (line, 6))
        diffusion_depth = read_cell(filepath, sheet, (line, 7))
        identification = read_cell(filepath, sheet, (line, 8))
        identification = clear_abnormal_characters(identification)
        identifications_with_preid = get_ID_with_preid(identification)
        for identification_with_preid in identifications_with_preid:
            products[-1].append(Product(year, month, day, heat_number, conpound_layer, diffusion_depth, identification_with_preid[0], identification_with_preid[1]))
        line += 1
    
    check_just(products)
    write_into_csv(products)

def check():
    """
    检查指定年份的 CSV 文件中是否有重复的 identification。
    """
    year = input("请输入需要检查的数据库的年份(如25): ")
    file_name = f"{year}database.csv"
    file_name = get_absolute_path(file_name)
    
    if not os.path.exists(file_name):
        print(f"文件 {file_name} 不存在。")
        return

    identification_dict = {}
    with open(file_name, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头
        for line_number, row in enumerate(reader, start=2):  # 从第二行开始，因为第一行是表头
            identification = int(row[7])
            if identification in identification_dict:
                identification_dict[identification].append(line_number)
            else:
                identification_dict[identification] = [line_number]

    duplicates_found = False
    for identification, lines in identification_dict.items():
        if len(lines) > 1:
            duplicates_found = True
            print(f"标识 {identification} 在以下行重复: {', '.join(map(str, lines))}")

    if not duplicates_found:
        print(f"文件 {file_name} 中没有重复的标识。")

def search():
    """
    交互式查询功能, 输入一个产品的完整编号, 如DL24-07698, 到对应的数据库中查询并返回结果。
    """
    full_identification = input("请输入产品的完整编号(如DL24-07698): ")
    search_product_in_csv(full_identification)
    

def generate_acceptance_report(list_of_identification: list[list[int]] = None):
    response = input("请输入日期(格式: 2025 1 21或者today):")
    if response == "today":
        year, month, day = time.localtime()[:3]
    else:
        year, month, day = map(int, response.split())
    if list_of_identification == None:
        list_of_identification = []
        response = input("请输入编号(以空格分隔, 附带编号前的年份, 比如24 07698 13456 25 19778):")
        list_of_identification_with_preid = get_ID_with_preid(response)
    table = ready_to_xlsx(list_of_identification_with_preid)
    fill_in(table, f"{year}_{month}_{day}_265缸套盐浴报告.xlsx")
    print("验收报告已生成")


def main():
    print("欢迎使用盐浴检查报告汇总文件便利程序, 请选择你要进行的操作:")
    print("1. 标准录入\n2. 信息检查与清洗\n3. 查询\n4. 生成验收报告")
    command = input("请输入你的选择: ")
    if command == "0":
        read_past_reports_and_write_into_csv()
    elif command == "1":
        standard_input()
    elif command == "2":
        check()
    elif command == "3":
        search()
    elif command == "4":
        generate_acceptance_report()
    else:
        print("输入有误, 请重新运行程序")

if __name__ == "__main__":
    main()